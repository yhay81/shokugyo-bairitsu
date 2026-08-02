"""Extract matched MHLW occupation openings and job-seeker tables for 2023–2025."""

from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl

YEARS = [2023, 2024, 2025]
OPENINGS_URL = "https://www.mhlw.go.jp/toukei/list/xls/114-1d-04.xlsx"
SEEKERS_URL = "https://www.mhlw.go.jp/toukei/list/xls/114-1d-05.xlsx"
SHEETS = {1: "a", 3: "f", 5: "t"}  # all, full-time excluding part, regular part-time
PREFECTURES = [
    ("JP-01", "北海道", "北海道"),
    ("JP-02", "青森", "東北"),
    ("JP-03", "岩手", "東北"),
    ("JP-04", "宮城", "東北"),
    ("JP-05", "秋田", "東北"),
    ("JP-06", "山形", "東北"),
    ("JP-07", "福島", "東北"),
    ("JP-08", "茨城", "関東"),
    ("JP-09", "栃木", "関東"),
    ("JP-10", "群馬", "関東"),
    ("JP-11", "埼玉", "関東"),
    ("JP-12", "千葉", "関東"),
    ("JP-13", "東京", "関東"),
    ("JP-14", "神奈川", "関東"),
    ("JP-15", "新潟", "北陸甲信越"),
    ("JP-16", "富山", "北陸甲信越"),
    ("JP-17", "石川", "北陸甲信越"),
    ("JP-18", "福井", "北陸甲信越"),
    ("JP-19", "山梨", "北陸甲信越"),
    ("JP-20", "長野", "北陸甲信越"),
    ("JP-21", "岐阜", "東海"),
    ("JP-22", "静岡", "東海"),
    ("JP-23", "愛知", "東海"),
    ("JP-24", "三重", "東海"),
    ("JP-25", "滋賀", "近畿"),
    ("JP-26", "京都", "近畿"),
    ("JP-27", "大阪", "近畿"),
    ("JP-28", "兵庫", "近畿"),
    ("JP-29", "奈良", "近畿"),
    ("JP-30", "和歌山", "近畿"),
    ("JP-31", "鳥取", "中国"),
    ("JP-32", "島根", "中国"),
    ("JP-33", "岡山", "中国"),
    ("JP-34", "広島", "中国"),
    ("JP-35", "山口", "中国"),
    ("JP-36", "徳島", "四国"),
    ("JP-37", "香川", "四国"),
    ("JP-38", "愛媛", "四国"),
    ("JP-39", "高知", "四国"),
    ("JP-40", "福岡", "九州・沖縄"),
    ("JP-41", "佐賀", "九州・沖縄"),
    ("JP-42", "長崎", "九州・沖縄"),
    ("JP-43", "熊本", "九州・沖縄"),
    ("JP-44", "大分", "九州・沖縄"),
    ("JP-45", "宮崎", "九州・沖縄"),
    ("JP-46", "鹿児島", "九州・沖縄"),
    ("JP-47", "沖縄", "九州・沖縄"),
]


def numeric(value: object) -> int | None:
    return int(value) if isinstance(value, (int, float)) and not isinstance(value, bool) else None


def load_openings(
    path: Path, place_ids: dict[str, str]
) -> tuple[dict, dict, dict, dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    values: dict[tuple[str, str, str], list[int | None]] = {}
    occupations: dict[str, dict[str, str]] = {}
    group_names: dict[str, str] = {}
    totals: dict[tuple[str, str], list[int | None]] = {}
    try:
        for sheet_index, employment in SHEETS.items():
            sheet = workbook.worksheets[sheet_index]
            current_place: str | None = None
            current_group: str | None = None
            for row in sheet.iter_rows(min_row=3, max_col=5, values_only=True):
                if row[0] in place_ids:
                    current_place = place_ids[str(row[0])]
                label = str(row[1] or "").strip()
                if current_place is not None and label == "職業計":
                    totals[(current_place, employment)] = [numeric(value) for value in row[2:5]]
                    continue
                group_match = re.match(r"^([Ａ-Ｋ])(.+)$", label)
                if group_match:
                    current_group = group_match.group(1)
                    group_names.setdefault(current_group, group_match.group(2))
                    continue
                occupation_match = re.match(r"^(\d{2})(.+)$", label)
                if current_place is None or current_group is None or not occupation_match:
                    continue
                occupation_id, occupation_name = occupation_match.groups()
                previous = occupations.setdefault(
                    occupation_id,
                    {"id": occupation_id, "name": occupation_name, "group": current_group},
                )
                if previous != {
                    "id": occupation_id,
                    "name": occupation_name,
                    "group": current_group,
                }:
                    raise ValueError(f"occupation changed across opening sheets: {label}")
                key = (current_place, occupation_id, employment)
                if key in values:
                    raise ValueError(f"duplicate opening series: {key}")
                values[key] = [numeric(value) for value in row[2:5]]
    finally:
        workbook.close()
    return values, occupations, group_names, totals


def load_seekers(
    path: Path,
    place_ids: dict[str, str],
    expected_occupations: dict[str, dict[str, str]],
) -> tuple[dict, dict, dict, int, int]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    values: dict[tuple[str, str, str], list[int | None]] = {}
    totals: dict[tuple[str, str], list[int | None]] = {}
    unclassified: dict[tuple[str, str], list[int | None]] = {}
    sex_checked = 0
    sex_mismatch = 0
    try:
        for sheet_index, employment in SHEETS.items():
            sheet = workbook.worksheets[sheet_index]
            current_place: str | None = None
            current_age: str | None = None
            current_group: str | None = None
            for row in sheet.iter_rows(min_row=4, max_col=12, values_only=True):
                if row[0] in place_ids:
                    current_place = place_ids[str(row[0])]
                if row[1] is not None:
                    current_age = str(row[1]).strip()
                if current_age != "年齢計" or current_place is None:
                    continue
                label = str(row[2] or "").strip()
                published = [numeric(row[offset]) for offset in (3, 6, 9)]
                if label == "職業計":
                    totals[(current_place, employment)] = published
                    continue
                if label == "分類不能の職業":
                    unclassified[(current_place, employment)] = published
                    continue
                group_match = re.match(r"^([Ａ-Ｋ])(.+)$", label)
                if group_match:
                    current_group = group_match.group(1)
                    continue
                occupation_match = re.match(r"^(\d{2})(.+)$", label)
                if not occupation_match or current_group is None:
                    continue
                occupation_id, occupation_name = occupation_match.groups()
                expected = expected_occupations.get(occupation_id)
                if expected != {
                    "id": occupation_id,
                    "name": occupation_name,
                    "group": current_group,
                }:
                    raise ValueError(f"occupation differs between sources: {label}")
                key = (current_place, occupation_id, employment)
                if key in values:
                    raise ValueError(f"duplicate seeker series: {key}")
                values[key] = published
                for offset in (3, 6, 9):
                    total, male, female = map(numeric, row[offset : offset + 3])
                    if total is not None and male is not None and female is not None:
                        sex_checked += 1
                        if total != male + female:
                            sex_mismatch += 1
    finally:
        workbook.close()
    return values, totals, unclassified, sex_checked, sex_mismatch


def main() -> None:
    if len(sys.argv) != 4:
        raise SystemExit(
            "usage: extract-source.py OPENINGS.xlsx SEEKERS.xlsx OUTPUT_DIRECTORY"
        )
    openings_path = Path(sys.argv[1])
    seekers_path = Path(sys.argv[2])
    output_directory = Path(sys.argv[3])
    places = [{"id": "JP-00", "name": "全国", "region": "全国"}] + [
        {"id": item_id, "name": name, "region": region}
        for item_id, name, region in PREFECTURES
    ]
    place_ids = {"全国計": "JP-00"} | {
        f"{name}労働局": item_id for item_id, name, _region in PREFECTURES
    }

    openings, occupations, group_names, opening_totals = load_openings(
        openings_path, place_ids
    )
    seekers, seeker_totals, unclassified, sex_checked, sex_mismatch = load_seekers(
        seekers_path, place_ids, occupations
    )
    if len(occupations) != 73 or len(group_names) != 11:
        raise ValueError(
            f"unexpected occupation dimensions: {len(occupations)} occupations, {len(group_names)} groups"
        )

    ordered_occupations = [occupations[key] for key in sorted(occupations)]
    records = []
    expected_series = len(places) * len(ordered_occupations) * len(SHEETS)
    if len(openings) != expected_series or len(seekers) != expected_series:
        raise ValueError(
            f"unexpected series count: openings={len(openings)} seekers={len(seekers)} expected={expected_series}"
        )
    for place in places:
        for occupation in ordered_occupations:
            record: dict[str, object] = {"p": place["id"], "o": occupation["id"]}
            for employment in SHEETS.values():
                opening_values = openings[(place["id"], occupation["id"], employment)]
                seeker_values = seekers[(place["id"], occupation["id"], employment)]
                record[employment] = [
                    [opening, seeker]
                    for opening, seeker in zip(opening_values, seeker_values, strict=True)
                ]
            records.append(record)

    identity_checked = {"openings": 0, "seekers": 0}
    for place in places:
        for occupation in ordered_occupations:
            for year_index in range(len(YEARS)):
                for label, source in (("openings", openings), ("seekers", seekers)):
                    all_value = source[(place["id"], occupation["id"], "a")][year_index]
                    full_value = source[(place["id"], occupation["id"], "f")][year_index]
                    part_value = source[(place["id"], occupation["id"], "t")][year_index]
                    if None not in (all_value, full_value, part_value):
                        identity_checked[label] += 1
                        if all_value != full_value + part_value:
                            raise ValueError(
                                f"employment identity mismatch: {label} {place['id']} {occupation['id']} {YEARS[year_index]}"
                            )

    national_sum_checked = {"openings": 0, "seekers": 0}
    prefecture_ids = [place["id"] for place in places if place["id"] != "JP-00"]
    for employment in SHEETS.values():
        for occupation in ordered_occupations:
            for year_index in range(len(YEARS)):
                for label, source in (("openings", openings), ("seekers", seekers)):
                    national = source[("JP-00", occupation["id"], employment)][year_index]
                    parts = [
                        source[(place_id, occupation["id"], employment)][year_index]
                        for place_id in prefecture_ids
                    ]
                    if national is not None and all(value is not None for value in parts):
                        national_sum_checked[label] += 1
                        if national != sum(parts):
                            raise ValueError(
                                f"national sum mismatch: {label} {employment} {occupation['id']} {YEARS[year_index]}"
                            )

    coverage = []
    for place in places:
        item: dict[str, object] = {"p": place["id"]}
        for employment in SHEETS.values():
            item[employment] = [
                [unknown, total]
                for unknown, total in zip(
                    unclassified[(place["id"], employment)],
                    seeker_totals[(place["id"], employment)],
                    strict=True,
                )
            ]
        coverage.append(item)

    pair_count = len(records) * len(SHEETS) * len(YEARS)
    available_source_values = 0
    calculable_ratios = 0
    unavailable_pairs = 0
    zero_denominators = 0
    zero_openings = 0
    for record in records:
        for employment in SHEETS.values():
            for opening, seeker in record[employment]:
                available_source_values += int(opening is not None) + int(seeker is not None)
                if opening is None or seeker is None:
                    unavailable_pairs += 1
                elif seeker == 0:
                    zero_denominators += 1
                else:
                    calculable_ratios += 1
                if opening == 0:
                    zero_openings += 1

    opening_sha = hashlib.sha256(openings_path.read_bytes()).hexdigest()
    seeker_sha = hashlib.sha256(seekers_path.read_bytes()).hexdigest()
    index = {
        "schemaVersion": 1,
        "asOf": "2026-08-02",
        "edition": "2023〜2025年度（現行表）",
        "years": YEARS,
        "placeCount": len(places),
        "prefectureCount": 47,
        "groupCount": len(group_names),
        "occupationCount": len(ordered_occupations),
        "recordCount": len(records),
        "employmentCount": len(SHEETS),
        "pairCount": pair_count,
        "sourceValueCount": pair_count * 2,
        "availableSourceValueCount": available_source_values,
        "unavailableSourceValueCount": pair_count * 2 - available_source_values,
        "calculableRatioCount": calculable_ratios,
        "unavailablePairCount": unavailable_pairs,
        "zeroDenominatorCount": zero_denominators,
        "zeroOpeningCount": zero_openings,
        "sexTotalChecked": sex_checked,
        "sexTotalMismatchCount": sex_mismatch,
        "employmentIdentityChecked": identity_checked,
        "nationalSumChecked": national_sum_checked,
        "places": places,
        "groups": [
            {"id": group_id, "name": group_names[group_id]}
            for group_id in sorted(group_names)
        ],
        "occupations": ordered_occupations,
        "coverage": coverage,
        "sources": [
            {
                "kind": "openings",
                "url": OPENINGS_URL,
                "bytes": openings_path.stat().st_size,
                "sha256": opening_sha,
            },
            {
                "kind": "seekers",
                "url": SEEKERS_URL,
                "bytes": seekers_path.stat().st_size,
                "sha256": seeker_sha,
            },
        ],
    }
    output_directory.mkdir(parents=True, exist_ok=True)
    (output_directory / "index.json").write_text(
        json.dumps(index, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    (output_directory / "ratios.json").write_text(
        json.dumps(records, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "calculable_ratios": calculable_ratios,
                "occupations": len(ordered_occupations),
                "pairs": pair_count,
                "places": len(places),
                "source_values": pair_count * 2,
                "unavailable_pairs": unavailable_pairs,
                "zero_denominators": zero_denominators,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
